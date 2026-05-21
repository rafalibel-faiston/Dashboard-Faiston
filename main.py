from fastapi import FastAPI, HTTPException, Response, Cookie
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
import psycopg2
import os, hashlib, secrets
from dotenv import load_dotenv
from pathlib import Path

load_dotenv()

app = FastAPI(title="Faiston Ops - API", version="1.0")

Path("static/css").mkdir(parents=True, exist_ok=True)
Path("static/js").mkdir(parents=True, exist_ok=True)

sessions = {}

def get_db():
    try:
        return psycopg2.connect(os.environ.get("DATABASE_URL"))
    except Exception as e:
        print(f"Erro BD: {e}")
        return None

def hash_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()

def get_session(token: str):
    return sessions.get(token)

def setup_banco():
    conn = get_db()
    if not conn: return
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id SERIAL PRIMARY KEY,
                usuario VARCHAR(50) UNIQUE NOT NULL,
                senha_hash VARCHAR(64) NOT NULL,
                nome VARCHAR(100) NOT NULL,
                perfil VARCHAR(20) NOT NULL DEFAULT 'funcionario',
                ativo BOOLEAN DEFAULT TRUE,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS tarefas (
                id SERIAL PRIMARY KEY,
                usuario_id INTEGER REFERENCES usuarios(id),
                descricao TEXT NOT NULL,
                cliente VARCHAR(50),
                prioridade VARCHAR(20) DEFAULT 'Media',
                status VARCHAR(30) DEFAULT 'aberto',
                segundos INTEGER DEFAULT 0,
                criado_em TIMESTAMP DEFAULT NOW(),
                atualizado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS comentarios (
                id SERIAL PRIMARY KEY,
                tarefa_id INTEGER REFERENCES tarefas(id) ON DELETE CASCADE,
                usuario_id INTEGER REFERENCES usuarios(id),
                texto TEXT NOT NULL,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS notificacoes (
                id SERIAL PRIMARY KEY,
                tipo VARCHAR(50) NOT NULL,
                mensagem TEXT NOT NULL,
                lida BOOLEAN DEFAULT FALSE,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("SELECT id FROM usuarios WHERE usuario = 'admin'")
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO usuarios (usuario, senha_hash, nome, perfil) VALUES (%s, %s, %s, %s)",
                ('admin', hash_senha('admin123'), 'Administrador', 'admin')
            )
        conn.commit(); cur.close(); conn.close()
        print("✅ Banco configurado")
    except Exception as e:
        print(f"Erro setup: {e}")

setup_banco()

# --- MODELOS ---
class LoginRequest(BaseModel):
    usuario: str
    senha: str

class NovoUsuario(BaseModel):
    usuario: str
    senha: str
    nome: str
    perfil: str

class AtualizarUsuario(BaseModel):
    nome: str
    perfil: str
    senha: str = ""
    ativo: bool = True

class TarefaModel(BaseModel):
    descricao: str
    cliente: str
    prioridade: str = "Media"
    status: str = "aberto"
    segundos: int = 0

class AtualizarSegundos(BaseModel):
    segundos: int

class AcaoBackoffice(BaseModel):
    comando: str
    cliente: str = "Geral"

# --- AUTH ---
@app.post("/api/login")
def login(req: LoginRequest, response: Response):
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, nome, perfil FROM usuarios WHERE usuario=%s AND senha_hash=%s AND ativo=TRUE",
                    (req.usuario, hash_senha(req.senha)))
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row: raise HTTPException(status_code=401, detail="Usuário ou senha inválidos")
        token = secrets.token_hex(32)
        sessions[token] = {"id": row[0], "nome": row[1], "perfil": row[2]}
        response.set_cookie("faiston_token", token, httponly=True, samesite="lax", max_age=86400)
        return {"sucesso": True, "perfil": row[2], "nome": row[1]}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/logout")
def logout(response: Response, faiston_token: str = Cookie(None)):
    if faiston_token and faiston_token in sessions: del sessions[faiston_token]
    response.delete_cookie("faiston_token")
    return {"sucesso": True}

@app.get("/api/me")
def me(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    return sess

# --- USUÁRIOS ---
@app.get("/api/usuarios")
def listar_usuarios(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor"): raise HTTPException(status_code=403, detail="Acesso negado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, usuario, nome, perfil, ativo, criado_em FROM usuarios ORDER BY criado_em DESC")
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id": r[0], "usuario": r[1], "nome": r[2], "perfil": r[3], "ativo": r[4], "criado_em": str(r[5])} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/usuarios")
def criar_usuario(u: NovoUsuario, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403, detail="Acesso negado")
    if u.perfil not in ("admin", "gestor", "funcionario"): raise HTTPException(status_code=400, detail="Perfil inválido")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO usuarios (usuario, senha_hash, nome, perfil) VALUES (%s, %s, %s, %s) RETURNING id",
                    (u.usuario, hash_senha(u.senha), u.nome, u.perfil))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except psycopg2.errors.UniqueViolation: raise HTTPException(status_code=400, detail="Usuário já existe")
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/usuarios/{uid}")
def atualizar_usuario(uid: int, u: AtualizarUsuario, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403, detail="Acesso negado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        if u.senha:
            cur.execute("UPDATE usuarios SET nome=%s, perfil=%s, ativo=%s, senha_hash=%s WHERE id=%s",
                        (u.nome, u.perfil, u.ativo, hash_senha(u.senha), uid))
        else:
            cur.execute("UPDATE usuarios SET nome=%s, perfil=%s, ativo=%s WHERE id=%s",
                        (u.nome, u.perfil, u.ativo, uid))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/usuarios/{uid}")
def deletar_usuario(uid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403, detail="Acesso negado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM usuarios WHERE id=%s", (uid,))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- TAREFAS ---
@app.get("/api/tarefas")
def listar_tarefas(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        if sess["perfil"] in ("admin", "gestor"):
            cur.execute("""SELECT t.id, t.descricao, t.cliente, t.prioridade, t.status, t.segundos, t.criado_em, u.nome
                FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id ORDER BY t.criado_em DESC""")
        else:
            cur.execute("""SELECT t.id, t.descricao, t.cliente, t.prioridade, t.status, t.segundos, t.criado_em, u.nome
                FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id
                WHERE t.usuario_id = %s ORDER BY t.criado_em DESC""", (sess["id"],))
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id": r[0], "descricao": r[1], "cliente": r[2], "prioridade": r[3],
                 "status": r[4], "segundos": r[5], "criado_em": str(r[6]), "funcionario": r[7]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/tarefas")
def criar_tarefa(t: TarefaModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO tarefas (usuario_id, descricao, cliente, prioridade, status, segundos) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id",
                    (sess["id"], t.descricao, t.cliente, t.prioridade, t.status, t.segundos))
        new_id = cur.fetchone()[0]
        criar_notificacao(conn, "nova_tarefa", f"🆕 {sess['nome']} criou uma tarefa: {t.descricao[:50]} [{t.cliente}]")
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/tarefas/{tid}")
def atualizar_tarefa(tid: int, t: TarefaModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("UPDATE tarefas SET descricao=%s, cliente=%s, prioridade=%s, status=%s, segundos=%s, atualizado_em=NOW() WHERE id=%s AND usuario_id=%s",
                    (t.descricao, t.cliente, t.prioridade, t.status, t.segundos, tid, sess["id"]))
        if t.status == "concluido":
            criar_notificacao(conn, "tarefa_concluida", f"✅ {sess['nome']} concluiu: {t.descricao[:50]} [{t.cliente}]")
        elif t.status == "em_andamento":
            criar_notificacao(conn, "tarefa_iniciada", f"▶️ {sess['nome']} iniciou: {t.descricao[:50]} [{t.cliente}]")
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.patch("/api/tarefas/{tid}/segundos")
def atualizar_segundos(tid: int, body: AtualizarSegundos, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("UPDATE tarefas SET segundos=%s, atualizado_em=NOW() WHERE id=%s AND usuario_id=%s",
                    (body.segundos, tid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/tarefas/{tid}")
def deletar_tarefa(tid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        if sess["perfil"] in ("admin", "gestor"):
            cur.execute("DELETE FROM tarefas WHERE id=%s", (tid,))
        else:
            cur.execute("DELETE FROM tarefas WHERE id=%s AND usuario_id=%s", (tid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- MÉTRICAS DASHBOARD ---
@app.get("/api/metricas")
def get_metricas(cliente: str = "", data_inicio: str = "", data_fim: str = "", funcionario: str = "", faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        conditions = []
        params = []
        if cliente:
            conditions.append("t.cliente = %s")
            params.append(cliente)
        if funcionario:
            conditions.append("u.nome ILIKE %s")
            params.append(f"%{funcionario}%")
        if data_inicio:
            conditions.append("t.criado_em >= %s")
            params.append(data_inicio + " 00:00:00")
        if data_fim:
            conditions.append("t.criado_em <= %s")
            params.append(data_fim + " 23:59:59")
        filtro = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        params = tuple(params)

        # JOIN usuarios necessário quando filtro por funcionário referencia u.nome
        join_u = "JOIN usuarios u ON t.usuario_id = u.id" if funcionario else ""
        # Helpers para adicionar condição de status sem quebrar o filtro existente
        def fwhere(extra): return f"{filtro} AND {extra}" if filtro else f"WHERE {extra}"

        # KPIs gerais
        cur.execute(f"SELECT COUNT(*) FROM tarefas t {join_u} {filtro}", params)
        total = cur.fetchone()[0]

        w_aberto = fwhere("t.status = 'aberto'")
        cur.execute(f"SELECT COUNT(*) FROM tarefas t {join_u} {w_aberto}", params)
        abertos = cur.fetchone()[0]

        w_concluido = fwhere("t.status = 'concluido'")
        cur.execute(f"SELECT COUNT(*) FROM tarefas t {join_u} {w_concluido}", params)
        concluidos = cur.fetchone()[0]

        w_andamento = fwhere("t.status = 'em_andamento'")
        cur.execute(f"SELECT COUNT(*) FROM tarefas t {join_u} {w_andamento}", params)
        em_andamento = cur.fetchone()[0]

        cur.execute(f"SELECT COALESCE(SUM(t.segundos),0) FROM tarefas t {join_u} {filtro}", params)
        total_segundos = cur.fetchone()[0]

        # Clientes ativos (distintos)
        cur.execute("SELECT COUNT(DISTINCT cliente) FROM tarefas")
        clientes_ativos = cur.fetchone()[0]

        # Funcionários com tarefas
        cur.execute("SELECT COUNT(DISTINCT usuario_id) FROM tarefas")
        funcionarios_ativos = cur.fetchone()[0]

        # SLA: % de tarefas concluídas sobre o total
        sla = round((concluidos / total * 100)) if total > 0 else 0

        # Média de horas por funcionário — respeita todos os filtros
        cur.execute(
            f"SELECT COUNT(DISTINCT t.usuario_id), COALESCE(SUM(t.segundos),0) FROM tarefas t {join_u} {filtro}",
            params
        )
        row_media = cur.fetchone()
        n_funcs = max(row_media[0], 1)
        media_horas_func = round(row_media[1] / 3600 / n_funcs, 1)

        # Horas por cliente — respeita todos os filtros
        cur.execute(
            f"SELECT t.cliente, COALESCE(SUM(t.segundos),0) as total_seg "
            f"FROM tarefas t {join_u} {filtro} GROUP BY t.cliente ORDER BY total_seg DESC",
            params
        )
        horas_por_cliente = [{"cliente": r[0], "horas": round(r[1]/3600, 1)} for r in cur.fetchall()]

        # Status da fila (para donut)
        cur.execute(f"SELECT t.status, COUNT(*) FROM tarefas t {join_u} {filtro} GROUP BY t.status", params)
        status_fila = {r[0]: r[1] for r in cur.fetchall()}

        # Volume por dia da semana (últimos 7 dias — para área)
        cur.execute("""
            SELECT TO_CHAR(criado_em, 'Dy') as dia, COUNT(*) as total
            FROM tarefas
            WHERE criado_em >= NOW() - INTERVAL '7 days'
            GROUP BY TO_CHAR(criado_em, 'Dy'), DATE_TRUNC('day', criado_em)
            ORDER BY DATE_TRUNC('day', criado_em)
        """)
        volume_semana = [{"dia": r[0], "total": r[1]} for r in cur.fetchall()]

        # Funil de atendimento
        funil = {
            "Abertura": total,
            "Triagem": total - max(0, total - abertos),
            "Acionamento": em_andamento + concluidos,
            "Acompanhamento": em_andamento + concluidos,
            "Fechamento": concluidos
        }

        # Tarefas recentes
        q = f"""SELECT t.id, t.descricao, t.cliente, t.prioridade, t.status, t.segundos, t.criado_em, u.nome
            FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id
            {filtro} ORDER BY t.criado_em DESC LIMIT 10"""
        cur.execute(q, params)
        recentes = [{"id": r[0], "descricao": r[1], "cliente": r[2], "prioridade": r[3],
                     "status": r[4], "segundos": r[5], "criado_em": str(r[6]), "funcionario": r[7]}
                    for r in cur.fetchall()]

        # Horas por funcionário — respeita todos os filtros
        func_conds = list(conditions) + ["u.perfil = 'funcionario'"]
        func_filtro = "WHERE " + " AND ".join(func_conds)
        cur.execute(
            f"SELECT u.nome, COALESCE(SUM(t.segundos),0), COUNT(t.id) as total_tarefas "
            f"FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id "
            f"{func_filtro} GROUP BY u.nome ORDER BY SUM(t.segundos) DESC",
            params
        )
        horas_por_func = [{"nome": r[0], "horas": round(r[1]/3600, 1), "tarefas": r[2]} for r in cur.fetchall()]

        # Taxa de conclusão por cliente — respeita todos os filtros
        cur.execute(
            f"SELECT t.cliente, COUNT(*) as total, "
            f"SUM(CASE WHEN t.status='concluido' THEN 1 ELSE 0 END) as concluidas "
            f"FROM tarefas t {join_u} {filtro} GROUP BY t.cliente ORDER BY total DESC",
            params)
        taxa_rows = cur.fetchall()
        taxa_conclusao = [{"cliente": r[0], "total": r[1], "concluidas": r[2],
            "taxa": round(r[2]/r[1]*100) if r[1] > 0 else 0} for r in taxa_rows]

        cur.close(); conn.close()
        return {
            "kpis": {
                "total_tarefas": total,
                "tickets_abertos": abertos,
                "em_andamento": em_andamento,
                "concluidos": concluidos,
                "clientes_ativos": clientes_ativos,
                "funcionarios_ativos": funcionarios_ativos,
                "total_horas": round(total_segundos / 3600, 1),
                "sla": sla,
                "media_horas_func": media_horas_func
            },
            "horas_por_cliente": horas_por_cliente,
            "status_fila": status_fila,
            "volume_semana": volume_semana,
            "funil": funil,
            "recentes": recentes,
            "horas_por_func": horas_por_func,
            "taxa_conclusao": taxa_conclusao
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/registrar-acao")
def registrar_acao(acao: AcaoBackoffice, faiston_token: str = Cookie(None)):
    return {"sucesso": True, "mensagem": "Ação registrada"}

@app.get("/api/health")
def health(): return {"status": "ok"}

@app.get("/api/exportar")
def exportar_excel(cliente: str = "", data_inicio: str = "", data_fim: str = "", faiston_token: str = Cookie(None)):
    from fastapi.responses import StreamingResponse
    import io
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    if sess["perfil"] not in ("admin", "gestor"): raise HTTPException(status_code=403, detail="Acesso negado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        conditions = []
        params = []
        if cliente:
            conditions.append("t.cliente = %s")
            params.append(cliente)
        if data_inicio:
            conditions.append("t.criado_em >= %s")
            params.append(data_inicio + " 00:00:00")
        if data_fim:
            conditions.append("t.criado_em <= %s")
            params.append(data_fim + " 23:59:59")
        filtro = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        cur.execute(f"""SELECT u.nome, t.descricao, t.cliente, t.prioridade, t.status,
            t.segundos, t.criado_em, t.atualizado_em
            FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id
            {filtro} ORDER BY t.criado_em DESC""", tuple(params))
        rows = cur.fetchall()
        cur.close(); conn.close()

        # Gerar XLSX com openpyxl
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tarefas"

        # Estilos
        header_fill = PatternFill("solid", fgColor="4A00E0")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill("solid", fgColor="F8FAFC")
        border = Border(bottom=Side(style='thin', color='E2E8F0'))
        center = Alignment(horizontal='center', vertical='center')

        # Cabeçalho
        headers = ["Funcionário", "Tarefa", "Cliente", "Prioridade", "Status", "Horas", "Minutos", "Total (h)", "Criado em", "Atualizado em"]
        col_widths = [25, 40, 20, 12, 15, 8, 8, 10, 18, 18]
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30

        # Dados
        status_map = {"concluido": "Concluído", "em_andamento": "Em Andamento", "aberto": "Aberto"}
        prio_colors = {"Alta": "FFE4E6", "Media": "FEF3C7", "Baixa": "D1FAE5"}
        status_colors = {"concluido": "D1FAE5", "em_andamento": "CFFAFE", "aberto": "F1F5F9"}

        for i, r in enumerate(rows, 2):
            h = r[5] // 3600
            m = (r[5] % 3600) // 60
            total_h = round(r[5] / 3600, 2)
            status_label = status_map.get(r[4], r[4])
            row_data = [r[0], r[1], r[2], r[3], status_label, h, m, total_h,
                str(r[6])[:16] if r[6] else "", str(r[7])[:16] if r[7] else ""]
            fill = PatternFill("solid", fgColor="FFFFFF") if i % 2 == 0 else alt_fill
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                # Cor por prioridade e status
                if col == 4 and r[3] in prio_colors:
                    cell.fill = PatternFill("solid", fgColor=prio_colors[r[3]])
                elif col == 5 and r[4] in status_colors:
                    cell.fill = PatternFill("solid", fgColor=status_colors[r[4]])
                else:
                    cell.fill = fill
            ws.row_dimensions[i].height = 22

        # Totais
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=sum(r[5]//3600 for r in rows)).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=round(sum(r[5] for r in rows)/3600, 2)).font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"faiston_tarefas{'_'+cliente if cliente else ''}{'_'+data_inicio if data_inicio else ''}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/seed-dados")
def seed_dados():
    """Rota temporária para popular banco com dados fictícios — delete após usar"""
    import random, hashlib
    from datetime import datetime, timedelta
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM tarefas")
        cur.execute("DELETE FROM usuarios WHERE usuario != 'admin'")
        funcionarios = [
            ("rafael","rafael123","Rafael Ribeiro Libel","funcionario"),
            ("pedro","pedro123","Pedro Alves","funcionario"),
            ("ezequiel","ezequiel123","Ezequiel Santos","funcionario"),
            ("ronald","ronald123","Ronald Ferreira","funcionario"),
            ("ana","ana123","Ana Carolina","funcionario"),
            ("lucas","lucas123","Lucas Martins","gestor"),
        ]
        ids = {}
        for usuario, senha, nome, perfil in funcionarios:
            cur.execute("INSERT INTO usuarios (usuario, senha_hash, nome, perfil) VALUES (%s,%s,%s,%s) ON CONFLICT (usuario) DO UPDATE SET nome=%s RETURNING id",
                (usuario, hashlib.sha256(senha.encode()).hexdigest(), nome, perfil, nome))
            ids[nome] = cur.fetchone()[0]
        clientes = ["NTT","Arcos Dourados","Zamp","Telcoweb"]
        prioridades = ["Alta","Alta","Media","Media","Media","Baixa"]
        status_opts = ["concluido","concluido","concluido","em_andamento","aberto"]
        descricoes = ["Abertura de chamado no NOC","Acompanhamento de incidente","Configuração de switch","Monitoramento de links MPLS","Troca de equipamento","Atualização de firmware","Relatório de disponibilidade","Escalada para fornecedor","Revisão de topologia","Acionamento de parceiro","Documentação de circuito","Teste de failover","Análise de log","Criação de tickets","Validação de SLA","Suporte remoto","Instalação de CPE","Diagnóstico de latência","Agendamento de manutenção","Follow-up de chamado crítico"]
        now = datetime.now()
        total = 0
        for nome, uid in ids.items():
            if nome == "Lucas Martins": continue
            for _ in range(random.randint(8,15)):
                criado_em = now - timedelta(days=random.randint(0,6), hours=random.randint(0,8), minutes=random.randint(0,59))
                status = random.choice(status_opts)
                segundos = random.randint(1800,14400) if status=="concluido" else random.randint(600,5400) if status=="em_andamento" else 0
                cur.execute("INSERT INTO tarefas (usuario_id,descricao,cliente,prioridade,status,segundos,criado_em,atualizado_em) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (uid, random.choice(descricoes), random.choice(clientes), random.choice(prioridades), status, segundos, criado_em, criado_em))
                total += 1
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "tarefas_criadas": total, "usuarios": [u[0] for u in funcionarios]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- PÁGINAS ---
@app.get("/")
def root(): return FileResponse("static/login.html")

@app.get("/dashboard")
def dashboard(): return FileResponse("static/index.html")

@app.get("/funcionario")
def funcionario(): return FileResponse("static/funcionario.html")

@app.get("/admin")
def admin_page(): return FileResponse("static/admin.html")

app.mount("/css", StaticFiles(directory="static/css"), name="css")
app.mount("/js", StaticFiles(directory="static/js"), name="js")

# --- COMENTÁRIOS ---
class ComentarioModel(BaseModel):
    texto: str

@app.get("/api/tarefas/{tid}/comentarios")
def listar_comentarios(tid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("""SELECT c.id, c.texto, c.criado_em, u.nome
            FROM comentarios c JOIN usuarios u ON c.usuario_id = u.id
            WHERE c.tarefa_id = %s ORDER BY c.criado_em ASC""", (tid,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "texto": r[1], "criado_em": str(r[2]), "autor": r[3]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/tarefas/{tid}/comentarios")
def criar_comentario(tid: int, c: ComentarioModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO comentarios (tarefa_id, usuario_id, texto) VALUES (%s,%s,%s) RETURNING id",
                    (tid, sess["id"], c.texto))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/comentarios/{cid}")
def deletar_comentario(cid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        if sess["perfil"] in ("admin", "gestor"):
            cur.execute("DELETE FROM comentarios WHERE id=%s", (cid,))
        else:
            cur.execute("DELETE FROM comentarios WHERE id=%s AND usuario_id=%s", (cid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- RELATÓRIO POR CLIENTE ---
@app.get("/api/relatorio/{cliente}")
def get_relatorio(cliente: str, mes: str = "", faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        params_base = [cliente]
        filtro_mes = ""
        if mes:
            filtro_mes = "AND DATE_TRUNC('month', t.criado_em) = DATE_TRUNC('month', %s::date)"
            params_base.append(mes + "-01")

        cur.execute(
            "SELECT t.id, t.descricao, t.prioridade, t.status, t.segundos, t.criado_em, t.atualizado_em, u.nome "
            "FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id "
            "WHERE t.cliente = %s " + filtro_mes + " ORDER BY t.criado_em DESC",
            params_base)
        tarefas = cur.fetchall()

        cur.execute(
            "SELECT u.nome, COUNT(t.id), COALESCE(SUM(t.segundos),0) "
            "FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id "
            "WHERE t.cliente = %s " + filtro_mes + " GROUP BY u.nome ORDER BY SUM(t.segundos) DESC",
            params_base)
        por_func = cur.fetchall()

        cur.execute(
            "SELECT status, COUNT(*) FROM tarefas t WHERE t.cliente = %s " + filtro_mes + " GROUP BY status",
            params_base)
        status_counts = {r[0]: r[1] for r in cur.fetchall()}

        cur.execute(
            "SELECT COALESCE(SUM(segundos),0) FROM tarefas t WHERE t.cliente = %s " + filtro_mes,
            params_base)
        total_seg = cur.fetchone()[0]

        cur.close(); conn.close()
        status_map = {"concluido": "Concluído", "em_andamento": "Em Andamento", "aberto": "Aberto"}
        return {
            "cliente": cliente, "mes": mes,
            "resumo": {
                "total_tarefas": len(tarefas),
                "concluidas": status_counts.get("concluido", 0),
                "em_andamento": status_counts.get("em_andamento", 0),
                "abertas": status_counts.get("aberto", 0),
                "total_horas": round(total_seg / 3600, 1),
                "sla": round(status_counts.get("concluido", 0) / len(tarefas) * 100) if tarefas else 0
            },
            "por_funcionario": [{"nome": r[0], "tarefas": r[1], "horas": round(r[2]/3600, 1)} for r in por_func],
            "tarefas": [{"id": r[0], "descricao": r[1], "prioridade": r[2],
                "status": status_map.get(r[3], r[3]), "horas": round(r[4]/3600, 1),
                "minutos": (r[4] % 3600) // 60, "criado_em": str(r[5])[:10],
                "funcionario": r[7]} for r in tarefas]
        }
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/relatorio/{cliente}")
def relatorio_page(cliente: str): return FileResponse("static/relatorio.html")

@app.get("/apresentacao")
def apresentacao_page(): return FileResponse("static/apresentacao.html")

# --- NOTIFICAÇÕES ---
def criar_notificacao(conn, tipo: str, mensagem: str):
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO notificacoes (tipo, mensagem) VALUES (%s, %s)", (tipo, mensagem))
        # Mantém só as últimas 50
        cur.execute("DELETE FROM notificacoes WHERE id NOT IN (SELECT id FROM notificacoes ORDER BY criado_em DESC LIMIT 50)")
        cur.close()
    except:
        pass

@app.get("/api/notificacoes")
def get_notificacoes(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    if sess["perfil"] not in ("admin", "gestor"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, tipo, mensagem, lida, criado_em FROM notificacoes ORDER BY criado_em DESC LIMIT 20")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "tipo": r[1], "mensagem": r[2], "lida": r[3], "criado_em": str(r[4])} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/notificacoes/marcar-lidas")
def marcar_lidas(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("UPDATE notificacoes SET lida = TRUE WHERE lida = FALSE")
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/notificacoes/nao-lidas")
def count_nao_lidas(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401)
    if sess["perfil"] not in ("admin", "gestor"): return {"count": 0}
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM notificacoes WHERE lida = FALSE")
        count = cur.fetchone()[0]
        cur.close(); conn.close()
        return {"count": count}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- HISTÓRICO COMPLETO ---
@app.get("/api/historico")
def get_historico(
    page: int = 1,
    por_pagina: int = 20,
    cliente: str = "",
    status: str = "",
    prioridade: str = "",
    funcionario: str = "",
    busca: str = "",
    data_inicio: str = "",
    data_fim: str = "",
    faiston_token: str = Cookie(None)
):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    if sess["perfil"] not in ("admin", "gestor"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        conditions = ["1=1"]
        params = []
        if cliente:
            conditions.append("t.cliente = %s")
            params.append(cliente)
        if status:
            conditions.append("t.status = %s")
            params.append(status)
        if prioridade:
            conditions.append("t.prioridade = %s")
            params.append(prioridade)
        if funcionario:
            conditions.append("u.nome ILIKE %s")
            params.append(f"%{funcionario}%")
        if busca:
            conditions.append("t.descricao ILIKE %s")
            params.append(f"%{busca}%")
        if data_inicio:
            conditions.append("t.criado_em >= %s")
            params.append(data_inicio + " 00:00:00")
        if data_fim:
            conditions.append("t.criado_em <= %s")
            params.append(data_fim + " 23:59:59")

        where = "WHERE " + " AND ".join(conditions)
        offset = (page - 1) * por_pagina

        # Total de registros
        cur.execute(f"SELECT COUNT(*) FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id {where}", params)
        total = cur.fetchone()[0]

        # Tarefas paginadas
        cur.execute(f"""
            SELECT t.id, t.descricao, t.cliente, t.prioridade, t.status,
                   t.segundos, t.criado_em, t.atualizado_em, u.nome
            FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id
            {where} ORDER BY t.criado_em DESC
            LIMIT %s OFFSET %s
        """, params + [por_pagina, offset])
        rows = cur.fetchall()

        # Resumo do filtro atual
        cur.execute(f"""
            SELECT COALESCE(SUM(t.segundos),0),
                   SUM(CASE WHEN t.status='concluido' THEN 1 ELSE 0 END),
                   COUNT(*)
            FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id {where}
        """, params)
        resumo = cur.fetchone()

        cur.close(); conn.close()
        status_map = {"concluido": "Concluído", "em_andamento": "Em Andamento", "aberto": "Aberto"}
        return {
            "total": total,
            "pagina": page,
            "por_pagina": por_pagina,
            "total_paginas": max(1, -(-total // por_pagina)),
            "resumo": {
                "total_horas": round((resumo[0] or 0) / 3600, 1),
                "concluidas": resumo[1] or 0,
                "total": resumo[2] or 0
            },
            "tarefas": [{
                "id": r[0], "descricao": r[1], "cliente": r[2],
                "prioridade": r[3], "status": status_map.get(r[4], r[4]),
                "segundos": r[5], "criado_em": str(r[6])[:16],
                "atualizado_em": str(r[7])[:16], "funcionario": r[8]
            } for r in rows]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/historico")
def historico_page(): return FileResponse("static/historico.html")

# --- NOTAS PESSOAIS ---
class NotaModel(BaseModel):
    titulo: str
    texto: str

@app.get("/api/notas")
def listar_notas(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS notas (
                id SERIAL PRIMARY KEY,
                usuario_id INTEGER REFERENCES usuarios(id) ON DELETE CASCADE,
                titulo VARCHAR(200) NOT NULL,
                texto TEXT,
                criado_em TIMESTAMP DEFAULT NOW(),
                atualizado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit()
        cur.execute("""SELECT id, titulo, texto, criado_em, atualizado_em
            FROM notas WHERE usuario_id = %s ORDER BY atualizado_em DESC""",
            (sess["id"],))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "titulo": r[1], "texto": r[2],
                 "criado_em": str(r[3])[:16], "atualizado_em": str(r[4])[:16]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/notas")
def criar_nota(n: NotaModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO notas (usuario_id, titulo, texto) VALUES (%s,%s,%s) RETURNING id",
                    (sess["id"], n.titulo, n.texto))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/notas/{nid}")
def atualizar_nota(nid: int, n: NotaModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("UPDATE notas SET titulo=%s, texto=%s, atualizado_em=NOW() WHERE id=%s AND usuario_id=%s",
                    (n.titulo, n.texto, nid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/notas/{nid}")
def deletar_nota(nid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM notas WHERE id=%s AND usuario_id=%s", (nid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- CLIENTES ---
@app.get("/api/clientes")
def listar_clientes(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS clientes (
                id SERIAL PRIMARY KEY,
                nome VARCHAR(100) UNIQUE NOT NULL,
                contato VARCHAR(100),
                email VARCHAR(100),
                ativo BOOLEAN DEFAULT TRUE,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        # Migrar clientes existentes das tarefas
        cur.execute("""
            INSERT INTO clientes (nome)
            SELECT DISTINCT cliente FROM tarefas
            WHERE cliente IS NOT NULL AND cliente != ''
            ON CONFLICT (nome) DO NOTHING
        """)
        conn.commit()
        cur.execute("SELECT id, nome, contato, email, ativo, criado_em FROM clientes WHERE ativo=TRUE ORDER BY nome")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "nome": r[1], "contato": r[2], "email": r[3], "ativo": r[4], "criado_em": str(r[5])[:10]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

class ClienteModel(BaseModel):
    nome: str
    contato: str = ""
    email: str = ""
    ativo: bool = True

@app.post("/api/clientes")
def criar_cliente(c: ClienteModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO clientes (nome, contato, email) VALUES (%s,%s,%s) RETURNING id",
                    (c.nome, c.contato, c.email))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except psycopg2.errors.UniqueViolation:
        raise HTTPException(status_code=400, detail="Cliente já existe")
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/clientes/{cid}")
def atualizar_cliente(cid: int, c: ClienteModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("UPDATE clientes SET nome=%s, contato=%s, email=%s, ativo=%s WHERE id=%s",
                    (c.nome, c.contato, c.email, c.ativo, cid))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/clientes/{cid}")
def deletar_cliente(cid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("UPDATE clientes SET ativo=FALSE WHERE id=%s", (cid,))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/clientes")
def clientes_page(): return FileResponse("static/clientes.html")